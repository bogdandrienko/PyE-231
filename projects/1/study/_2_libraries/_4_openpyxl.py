import json
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import time


# Read
# Create
# Change (Read + Create)
# Template (уже есть красивый файл с внешним видом / форматированием)

def read():
    workbook: Workbook = openpyxl.load_workbook("temp/data.xlsx")
    worksheet: Worksheet = workbook.active  # todo АКТИВИРУЕТ АКТИВНЫЙ
    # worksheet = workbook["Рабочий лист"]  # todo АКТИВИРУЕТ нужный рабочий лист

    # list1: list[int] = [1, 2, 3]
    # list1: list[str] = ["1"]
    # list1: list[any] = ["1", 2, 3]

    cell1 = worksheet.cell(row=2, column=1).value
    print(cell1, type(cell1))

    cell2 = worksheet["A3"].value
    print(cell2, type(cell2))

    matrix = []
    for row_i in range(1, 5 + 1):  # 1 2 3 4
        local_list = []
        for column_i in range(1, 3 + 1):  # 1 2
            # 1*1
            # 1*2

            # 2*1
            # 2*2

            # 3*1
            # 3*2

            # 4*1
            # 4*2
            value = worksheet.cell(row=row_i, column=column_i).value
            local_list.append(value)
        matrix.append(local_list)

    print(matrix)
    print("\n\n\n\n")

    for j in matrix:
        print(j)
    pass


def practic():
    # прочитать excel
    workbook: Workbook = openpyxl.load_workbook("temp/work.xlsx")
    worksheet: Worksheet = workbook.active

    matrix = []
    for row_i in range(1, worksheet.max_row + 1):
        local_list = []
        for column_i in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=row_i, column=column_i).value
            local_list.append(value)
        matrix.append(local_list)
    # for j in matrix:
    #     print(j)

    # прочитать txt
    with open("temp/equal.txt", "r", encoding="utf-8") as file:
        lines = file.readlines()
    # print(type(lines), lines)
    clear_lines = []
    for i in lines:
        clear_lines.append(int(i))
    # print(type(clear_lines), clear_lines)
    # print(184 in clear_lines)

    # сравнить
    result = []
    for j in matrix:
        if j[0] in clear_lines:
            new_dict = {"value":  j[0], "stepen": j[1]}
            result.append(new_dict)

    # записать итоговые данные
    with open("temp/equal.json", "w", encoding="utf-8") as file:
        json.dump(result, file)


if __name__ == "__main__":
    # read()
    practic()
