# Практика: https://jsonplaceholder.typicode.com/todos/N
# 1. загрузить и сохранить в файл одну "тудушку"
# 2. загрузить последовательно 10 файлов, и замерить производительность (декоратор)
# 3. загрузить параллельно (threading) 10 файлов, и замерить производительность (декоратор)
# 4. загрузить параллельно (multiprocessing) 10 файлов, и замерить производительность (декоратор)
# https://github.com/bogdandrienko/PyE-231/blob/main/projects/1/study/_1_base/_15_threading_processing_async.py
import multiprocessing
import threading
import time
import requests
import json


def decorator_time_measure(func):
    def wrapper(*args, **kwargs):
        time_start = time.perf_counter()

        result = func(*args, **kwargs)

        print("elapsed_time(s): ", round(time.perf_counter() - time_start, 6))
        return result

    return wrapper


def get_one_todo(num: int):
    url = f"https://jsonplaceholder.typicode.com/todos/{num}"
    response = requests.get(url)
    data = response.json()

    with open(f"temp/data{num}.json", mode="w", encoding="utf-8") as file:
        json.dump(data, file)


@decorator_time_measure
def get_10_todo_sync():
    for i in range(1, 10 + 1):
        get_one_todo(i)


@decorator_time_measure
def get_10_todo_threading():
    thread_list = []
    for i in range(1, 10 + 1):
        thread_list.append(threading.Thread(target=get_one_todo, args=(i,), kwargs={}))
    for thread in thread_list:
        thread.start()
    for thread in thread_list:
        thread.join()


@decorator_time_measure
def get_10_todo_multiprocessing():
    process_list = []
    for i in range(1, 10 + 1):
        process_list.append(multiprocessing.Process(target=get_one_todo, args=(i,), kwargs={}))
    for process in process_list:
        process.start()
    for process in process_list:
        process.join()


if __name__ == "__main__":
    # get_one_todo(66)
    # get_10_todo_sync()                # 2.879422
    # get_10_todo_threading()           # 0.549079
    # get_10_todo_multiprocessing()     # 0.761359
    pass

def extra():
    from openpyxl import Workbook, load_workbook

    col1 = []
    workbook = load_workbook('Лист1.xlsx')
    sheet = workbook.active
    for i in range(1, sheet.max_row + 1):
        col1.append(sheet.cell(row=i, column=1))

    col2 = []
    workbook = load_workbook('Лист2.xlsx')
    sheet = workbook.active
    for i in range(1, sheet.max_row + 1):
        col1.append(sheet.cell(row=i, column=2))

    col3 = []
    workbook = load_workbook('Лист2.xlsx')
    sheet = workbook.active
    for i in range(1, sheet.max_row + 1):
        col1.append(sheet.cell(row=i, column=3))

    matrix = [col1, col2, col3]

    combined_file = Workbook()
    combined_sheet = combined_file.active
    row_i = 0
    for i in matrix:
        row_i += 1
        col_i = 0
        for j in i:
            col_i += 1
            combined_sheet.cell(row=row_i, column=col_i, value=j)
    combined_file.save('combined_file.xlsx')