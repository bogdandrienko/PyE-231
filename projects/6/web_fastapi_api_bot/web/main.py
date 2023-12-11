import datetime
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import io
import sqlite3
# async
# import aioredis
import aiosqlite
import aiohttp
import aiofiles
from starlette.responses import RedirectResponse
from fastapi.templating import Jinja2Templates
from fastapi import FastAPI, Request, Depends

app = FastAPI()
templates = Jinja2Templates(directory="templates")


@app.get("/")
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.post("/submit")
async def submit(request: Request):
    # form = await request.json()
    # input_text = form_data.get("input_text")
    # print(form)
    form_data = await request.form()  # пока форма грузится, можно жарить яйца
    file = form_data["addition"]
    f = await file.read()  # пока форма грузится, можно жарить яйца
    # print(type(await file.read()), await file.read())
    workbook: Workbook = openpyxl.load_workbook(filename=io.BytesIO(f), data_only=True)
    worksheet: Worksheet = workbook.active
    data = tuple(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column, values_only=True))
    print(data)
    clear_data = []
    for i in data:
        is_have_none = False
        for j in i:
            if j is None:
                is_have_none = True
        if not is_have_none:
            clear_data.append(i)
        # if all(j is not None for j in i):  # синтаксический сахар
        #     clear_data.append(i)
    print(clear_data)


    async with aiosqlite.connect('database.db') as db:
        await db.execute('''
CREATE TABLE IF NOT EXISTS items (
    id INTEGER PRIMARY KEY UNIQUE NOT NULL,
    name TEXT NOT NULL,
    description TEXT NOT NULL,
    price REAL default 0.0,
    quantity INTEGER default 0
)''')  # пока форма грузится, можно жарить яйца
        await db.commit()  # пока форма грузится, можно жарить яйца

        # SQL Injection - уязвимость в БД
        # пришли данные, из файла, из input(html form)...
        # злоумышленник/вредитель мог вставит в поле вредоносный SQL код
#         await db.execute(f'''
# INSERT INTO my_table (id, name, description, price, quantity)
# VALUES ({data[0]}, 'Боты', 'Удобные боты', 70.5, 3)
# ''')
        it = clear_data[0]
        beaty = f"[{it[0]}] {it[1]} - {it[2][:30]} ... {it[3]}$ | {it[4]}"
        print(it)
        # INSERT OR REPLACE INTO items (id, name, description...)
        # UPSERT(UPDATE OR INSERT)
        await db.execute('''
INSERT INTO items (id, name, description, price, quantity)
VALUES (:id, :name, :description, :price, :quantity)
''', {"id": it[0], "name": it[1], "description": it[2], "price": it[3], "quantity": it[4]})  # todo sql injection SAFE
        # placeholder - держатель места (ЗАГЛУШКА)
        await db.commit()  # пока форма грузится, можно жарить яйца

    bot_token = "6669581355:AAEcK9yBwL_D2tZi8W6zUVup04STZYkStIE"
    users = '1289279426,431390376,418009489,795567664'  #1289279426(bogdandrienko), BladEugene 431390376, supernaturalmlady 418009489, Dossanov_Baizhan 795567664
    async with aiohttp.ClientSession() as session:
        for user in [str(x).strip() for x in users.split(',')]:
            async with session.get(f'https://api.telegram.org/bot{bot_token}/sendMessage', params={'chat_id': user, 'text': beaty}) as response:
                await response.json()  # пока форма грузится, можно жарить яйца
            # try:
            #     async with session.get(f'https://api.telegram.org/bot{bot_token}/sendMessage', params={'chat_id': user, 'text': beaty}) as response:
            #         response = await response.json()
            #         if response.status_code not in (200, 201):  # 200 - ok, 201 - created
            #             raise Exception(response.status_code)
            # except Exception as error:
            #     print(error)
            #     async with session.get(f'https://api.telegram.org/bot{bot_token}/sendMessage', params={'chat_id': user, 'text': beaty}) as response:
            #         response = await response.json()
            #         if response.status_code not in (200, 201):
            #             async with aiofiles.open("logs.txt", 'a', encoding="utf-8") as f:
            #                 await f.write(f"{datetime.datetime.now()} {error}")

    return RedirectResponse(url="/")  # GET
