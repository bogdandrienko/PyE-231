import sys
import sqlite3
import datetime
import threading
import time
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PyQt6.QtWidgets import QApplication, QWidget, QMessageBox, QGridLayout, QPushButton
from PyQt6 import uic
import smtplib as smtp
from getpass import getpass

# сборка в exe: auto-py-to-exe


class Ui(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = uic.loadUi("main.ui", self)
        self.ui.pushButton_save.clicked.connect(self.save_to_database)
        self.ui.pushButton_export.clicked.connect(self.export_from_database)
        self.ui.pushButton_send_mail.clicked.connect(self.send_mail)

        self.ui.label_total_count.setText("-")

        self.show()
        self.update_loop()

    def update_loop(self):
        def update():
            try:
                while True:
                    self.update_total_count_items()
                    time.sleep(10)
            except Exception as error:
                print(error)

        new_thread = threading.Thread(target=update, args=(), kwargs={})
        new_thread.start()

    def save_to_database(self):
        try:
            # номер заявки
            number = self.lineEdit_number.text().strip()
            is_number = self.validate_number(value=number)
            if is_number is False:
                # self.app2 = Modal()  # всплывающее модальное окно
                # self.app2.show()
                QMessageBox.critical(self, "Ошибка", "Вы ввели неправильный номер заявки")
                return

            # сумма
            price = self.doubleSpinBox_price.value()
            # is_price = self.validate_price(value=price)
            # if is_number is False:
            #     QMessageBox.critical(self, "Ошибка", "Вы ввели неправильный сумму")
            #     return

            # TODO Запись в базу данных
            # print(number, price)

            # отправка в базу данных
            connection = sqlite3.connect("database_items.db")
            cursor = connection.cursor()

            # TODO SQL INJECTION ########################################
            # вредоносный SQL-код
            # number = '124314;drop table postgres;'
            query = f"""
INSERT INTO items (number, price)
VALUES ('{number}', '{price}')
"""
            # TODO SQL INJECTION ########################################

            # safe to SQL INJECTION
            # сама библиотека проверяет все переменные и удаляет возможный вредоносный код
            query = f"""
            INSERT INTO items (number, price)
            VALUES (?, ?)
            """
            # oracle ':param_date'  postgres '%s'
            cursor.execute(query, (number, price))
            connection.commit()

            self.update_total_count_items()

            QMessageBox.information(self, "Успешно", "Заявка успешно добавлена в базу данных")
        except Exception as error:
            print(error)

    def export_from_database(self):
        try:
            # print("Экспорт из базы данных")
            # получение всех данных из базы данных
            rows: list[tuple[any]] = self.get_all_items()
            # TODO МОЖЕТ ЗАНЯТЬ ДО 10 секунд

            # [
            # (1, 0, 666.67),
            # (2, 134000, 666.67),
            # (3, 12345, 12.0),
            # (4, 54321, 33.0),
            # (5, 5555555, 55.0)
            # ]

            # создание в оперативной памяти excel файла и его открытие
            workbook: Workbook = openpyxl.Workbook()
            worksheet: Worksheet = workbook.active

            # запись заголовков в excel-файл
            for column_i, column in enumerate(["ИИН", "Номер заявки", "Сумма заявки"], 1):
                worksheet.cell(row=1, column=column_i, value=column)

            # запись всех данных в excel-файл
            for row_i, row in enumerate(rows, 2):
                for column_i, column in enumerate(row, 1):
                    worksheet.cell(row=row_i, column=column_i, value=column)

            workbook.save(f"экспорт_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx")
            QMessageBox.information(self, "Успешно", "Данные успешно экспортированы")
        except Exception as error:
            print(error)

    def validate_number(self, value: str) -> bool:
        """Проверка номера заявки число"""
        try:
            value = int(value)
            if value < 1:
                return False
        except Exception as error:
            print("Error: ", error)
            return False
        return True

    def validate_price(self):
        pass

    def get_all_items(self):
        connection = sqlite3.connect("database_items.db")
        cursor = connection.cursor()
        query1 = """
SELECT * FROM items;
"""
        query2 = """
INSERT INTO items (number, price)
VALUES ('134000', '666.67')
"""
        query3 = """
SELECT COUNT(*) FROM items;
"""
        cursor.execute(query1)
        # connection.commit()
        rows = cursor.fetchall()  # fetchone()
        print(rows)
        return rows

    def update_total_count_items(self):
        # если действие в один поток
        # а база данных медленная, занята... и времени уйдёт 3 с.
        def update():
            try:
                # time.sleep(3.0)
                # запрос к базе данных
                count = query_to_sqlite(_query="""SELECT count(*) FROM items;""", many=False)[0]
                self.ui.label_total_count.setText(f"{count}")
            except Exception as error:
                error_mes = str(error)
                with open("log.txt", "a") as file:
                    file.write(f"{error_mes} {datetime.datetime.now()}\n")
                print(error_mes)
                self.ui.label_total_count.setText(error_mes)

        new_thread = threading.Thread(target=update, args=(), kwargs={})
        new_thread.start()

    def send_mail(self):
        email_from = "eevee.cycle@yandex.com"
        password = "mawfgptgdehfcwyi"
        email_to = "bogdandrienko@gmail.com"
        subject = "Python send mail"

        count = query_to_sqlite(_query="""SELECT count(*) FROM items;""", many=False)[0]
        text = f"Items count: {count}."

        # format1 = "From: {}\nTo: {}\nSubject: {}\n\n{}".format(email_from, email_to, subject, text)
        # format2 = "From: %s\nTo: %s\nSubject: %s\n\n%s" % (email_from, email_to, subject, text)
        message = f"From: {email_from}\nTo: {email_to}\nSubject: {subject}\n\n{text}"

        server = smtp.SMTP_SSL("smtp.yandex.com:465")  # SSL = 465 | TLS = 587
        # server.set_debuglevel(1)
        # server.starttls()
        server.login(email_from, password)
        server.sendmail(email_from, email_to, message)
        server.quit()


def query_to_sqlite(_query: str, many: bool, _args: tuple = (), _source: str = "database_items.db"):
    with sqlite3.connect(_source) as connection:
        # with connection.cursor() as cursor:  # TODO POSTGRESQL
        cursor = connection.cursor()
        cursor.execute(_query, _args)
        if many:
            return cursor.fetchall()  # todo list[tuple[any]]
        return cursor.fetchone()  # todo tuple[any]


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ui = Ui()
    sys.exit(app.exec())

    def database_postgre():
        """
        CREATE TABLE items (
        id SERIAL PRIMARY KEY,
        number BIGINT NOT NULL,
        price DECIMAL(10, 2) NOT NULL default = '0.0'
        );
        """

    def database_sqlite():
        # создание соединения с базой данных
        connection = sqlite3.connect("database_items.db")  # SELECT name FROM sqlite_master WHERE type='table';

        # создание "курсора" к базе данных
        cursor = connection.cursor()

        # Написание запроса для создания базы данных на SQL
        query = """
CREATE TABLE IF NOT EXISTS items 
(
id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
number INTEGER NOT NULL,
price REAL default '0.0'
)
"""
        # Выполнение запроса
        cursor.execute(query)

    class Modal(QWidget):
        def __init__(self, parent=None):
            super().__init__(parent)
            # self.setWindowModality(Qt.WindowModal)
            # self.setModal(True)
            self.resize(400, 400)
            self.setWindowTitle("Вы ввели неверный номер заявки")

            self.grid = QGridLayout(self)

            self.button = QPushButton()
            self.button.setText("ОК")

            self.grid.addWidget(self.button)
