import random

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# salary_s = [70000, 80000, 90000]
# position_s = ["Программист", "Инженер", "Сис.админ"]
# name_s = ["Вася", "Петя", "Катя"]
#
# with open("workers.txt", "a", encoding="utf-8") as file:
#     for _ in range(1, 10000):
#         file.write(f"{random.choice(name_s)}, {random.choice(position_s)} = {random.choice(salary_s)}\n")

# сложить всех работников по категориям, получить среднее и медиану
# 1 * 10
# 6 * 4
# (10 + 4) / 2 = 7
# 34 / 7 = 4.85

"""
1. Прочитать файл
2. Сложить по категориям
"""


class Worker:
    ADDITIONAL_BONUSES = 50000
    # _
    # __

    def __repr__(self):
        return f"<Worker {self.name} {self.position} {self.salary}>"

    def __init__(self, name: str, position: str, salary: int | float):
        self.name = name
        self.position = position
        self.salary = salary

    def get_clear_salary(self):  # динамический класс(есть self - ссылка на себя)
        if self.position.startswith("Ведущий"):
            return self.salary - self.ADDITIONAL_BONUSES
        return self.salary


workers: list[Worker] = []
with open("workers.txt", "r", encoding="utf-8") as file:
    lines = file.readlines()
    # print(type(lines), lines)
    for line in lines:
        try:
            clear_line = line.strip()
            name = str(clear_line.split(",")[0].strip())
            position = str(clear_line.split(",")[1].split("=")[0].strip())
            salary = float(clear_line.split("=")[1].strip())
            # 'Катя, Сис.админ = 70000\n'
            # print(name, position, salary)
            new_worker = Worker(name=name, position=position, salary=salary)
            workers.append(new_worker)
        except Exception as error:
            print("Error: ", error)

print(type(workers))  # <class 'list'>
print(type(workers[0]))  # <class '__main__.Worker'>

summary = 0
summ_sys = 0
for worker in workers:
    summary += worker.salary
    if worker.position == "Сис.админ":
        summ_sys += worker.salary
print(summary)
print(summ_sys)


def for_filter(w) -> bool:
    res: bool = w.position == "Сис.админ"
    return res


# list2 = list(filter(lambda w: w.position == "Сис.админ", workers))
list2: list[Worker] = list(filter(for_filter, workers))
salarys = sum([x.salary for x in list2])
print(salarys, len(list2), salarys / len(list2), list2)

"""
Посчитать сумму и среднюю зарплат в каждой категории
"""

new_workbook = openpyxl.Workbook()
new_worksheet: Worksheet = new_workbook.active
index = 0
for i in workers:
    index += 1
    new_worksheet.cell(row=index, column=1, value=i.name)
    new_worksheet.cell(row=index, column=2, value=i.position)
    new_worksheet.cell(row=index, column=3, value=i.salary)
new_workbook.save("data.xlsx")



