"""
json -> txt/Excel file/...
"""
import os
import random

import requests
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

response = requests.get("https://jsonplaceholder.typicode.com/posts")
data: list[dict] = response.json()
print(type(data), len(data))
for i in data:
    print(i)
#
# new_workbook: Workbook = openpyxl.Workbook()
# new_worksheet: Worksheet = new_workbook.active
#
# # запись заголовков в excel-файл
# titles = ["userId", "id", "title", "body"]
# index = 0
# for title in titles:
#     index += 1
#     new_worksheet.cell(row=1, column=index, value=title)
#
# index = 1
# for d in data:
#     index += 1
#     new_worksheet.cell(row=index, column=1, value=d["userId"])
#     new_worksheet.cell(row=index, column=2, value=d["id"])
#     new_worksheet.cell(row=index, column=3, value=d["title"])
#     new_worksheet.cell(row=index, column=4, value=d["body"])
#
# new_workbook.save("new_data.xlsx")

"""
-> list[list[any]], 'data.xlsx'(O/I)
-> bool
"""


class WorkWithExcel:
    @staticmethod
    def get_workbook_if_exists(filename: str, rewrite: bool) -> Workbook:
        # 1. Проверить, существует ли файл?
        # 2. Попытаться прочитать файл, если его нет, то создать
        try:
            # todo - возможно, стоит при ошибке удаления файла(файл занят), генерить другое имя
            if rewrite:
                os.remove(filename)
            workbook = openpyxl.load_workbook(filename)
        except Exception as _:
            workbook = openpyxl.Workbook()
        return workbook

    @staticmethod
    def convert_list_dict_to_matrix(list_dict: list[dict]) -> list[list]:
        """
        [{"id": 1, "name": "Bogdan"},]
        [[1, "Bogdan"],]
        """
        global_result = []
        for dicts in list_dict:
            # dicts.items()  # ["name", "Bogdan"]
            # dicts.keys() ["id", "name"]
            # dicts.values() [1, "Bogdan"]
            local_result = [*dicts.values()]
            global_result.append(local_result)
        return global_result

    @staticmethod
    def write_to_excel(filename: str, matrix: list[list] | list[dict], titles: list[str], rewrite=True) -> bool:
        try:
            if len(matrix) < 1:
                raise Exception("Length is lower than 1!")
            workbook = WorkWithExcel.get_workbook_if_exists(filename=filename, rewrite=rewrite)
            if isinstance(matrix[0], dict):
                matrix = WorkWithExcel.convert_list_dict_to_matrix(matrix)
            # print(type(matrix), type(matrix[0]), matrix)

            worksheet: Worksheet = workbook.active

            index = 0
            for title in titles:
                index += 1
                worksheet.cell(row=1, column=index, value=title)

            index_1 = 1
            for local_matrix in matrix:
                index_1 += 1
                index_2 = 0
                for value in local_matrix:
                    index_2 += 1
                    # index_1 = строки
                    # index_2 = колонки
                    worksheet.cell(row=index_1, column=index_2, value=value)
            workbook.save(filename)

            return True
        except Exception as error:
            print("Error: ", error)
            return False


mat = [
    [1, 2, 3, 4, 5],
    [5, 4, 3, 2, 1],
    [1, 2, 3, 4, 5],
    [5, 4, 3, 2, 1],
]

mat2 = [
    {"id": 1, "name": "Bogdan"},
    {"id": 2, "name": "Bogdan 2"},
    {"id": 3, "name": "Bogdan 3"},
]

_titles: list[str] = ["id", "name", "body"]

WorkWithExcel.write_to_excel(filename="new_file_data.xlsx", matrix=data, titles=_titles)


class Worker:  # (object)
    salary = 777  # атрибут класса

    def __init__(self, position: str):  # магический метод(функция, внутри класса) -
        self.position = position  # атрибут экземпляра класса


class Stage(Worker):
    def __init__(self, position: str):
        super().__init__(position)
        self.salary = super().salary / 2


salary_s = [70000, 80000, 90000]
position_s = ["Программист", "Инженер", "Сис.админ"]
name_s = ["Вася", "Петя", "Катя"]

with open("workers.txt", "a") as file:
    for _ in range(1, 10000):
        file.write(f"{random.choice(name_s)}, {random.choice(position_s)} = {random.choice(salary_s)}")
