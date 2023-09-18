import datetime
import json
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class Person:
    """
    1. Реализовать класс для заданной структуры.
    2. Прочитать весь файл в оперативную память, в эту структуру.
    3. Реализовать возврат каждой сущности в json.
    4. Отфильтровать только людей в кадровом резерве.
    5. Отсортировать в порядке возрастания.
    6. Добавить им всем БОНУС + 15% к зарплате и выгрузить их.

    Имя	Должность	Зарплата	Кадровый резерв	Дата устройства
    Вася	Инженер	90000	да	1 января 2023 г.
    """

    PERSONAL_BONUS = 1.15

    def __repr__(self):
        if self.is_reserved:
            s = "В резерве"
        else:
            s = "Не в резерве"
        return f"<Person {self.name} {self.salary} {s} {self.data_registered}>"

    def __init__(self, name: str, position: str, salary: any, is_reserved: str, data_registered):
        self.name: str = name
        self.position: str = position
        self.salary: float = float(salary)
        self.is_reserved: bool = str(is_reserved).strip().lower() == "да"
        self.data_registered: datetime.datetime = data_registered

    def serialize_to_json(self) -> dict:
        return {"name": self.name, "position": self.position, "salary": self.salary,
                "is_reserved": self.is_reserved, "data_registered": str(self.data_registered)}

    @staticmethod
    def write_list_to_json(data: list[dict]):
        with open("data.json", "w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False)

    @staticmethod
    def write_list_to_txt(data: list[dict]):
        with open("data.txt", "w", encoding="utf-8") as file:
            for d in data:
                file.write(str(d) + "\n")


persons: list[Person] = []
workbook: Workbook = openpyxl.load_workbook("data.xlsx")
worksheet: Worksheet = workbook.active
for i in range(2, worksheet.max_row + 1):
    person = Person(
        name=worksheet.cell(row=i, column=1).value,
        position=worksheet.cell(row=i, column=2).value,
        salary=worksheet.cell(row=i, column=3).value,
        is_reserved=worksheet.cell(row=i, column=4).value,
        data_registered=worksheet.cell(row=i, column=5).value,
    )
    persons.append(person)
print(persons)

# отсев тех, кто не в резерве
only_reserved: list[Person] = list(filter(lambda x: x.is_reserved, persons))
print(only_reserved)

# сортировка по дате регистрации
only_reserved_sorted = sorted(only_reserved, key=lambda p: p.data_registered, reverse=False)
print(only_reserved_sorted)

# взятие первых четверых
first_4 = only_reserved_sorted[:4]
print(first_4)

# конвертация из объектов в словари
data: list[dict] = []
for p in first_4:
    p.salary = p.salary * p.PERSONAL_BONUS
    data.append(p.serialize_to_json())

# запись в json-файл
Person.write_list_to_json(data)
Person.write_list_to_txt(data)
