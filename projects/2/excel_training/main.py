import json
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def practic():
    # TODO прочитать excel
    workbook: Workbook = openpyxl.load_workbook("temp/work.xlsx")
    worksheet: Worksheet = workbook.active

    matrix = []
    for row_i in range(1, worksheet.max_row + 1):
        local_list = []
        for column_i in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=row_i, column=column_i).value
            local_list.append(value)
        matrix.append(local_list)
    # for j in matrix:
    #     print(j)

    # TODO прочитать txt
    with open("temp/equal.txt", "r", encoding="utf-8") as file:
        lines = file.readlines()
    # print(type(lines), lines)
    clear_lines = []
    for i in lines:
        clear_lines.append(int(i))
    # print(type(clear_lines), clear_lines)
    # print(184 in clear_lines)

    # TODO сравнить
    result = []
    for j in matrix:
        if j[0] in clear_lines:
            new_dict = {"value": j[0], "stepen": j[1]}
            result.append(new_dict)

    # TODO записать итоговые данные
    with open("temp/equal.json", "w", encoding="utf-8") as file:
        json.dump(result, file)


if __name__ == "__main__":
    practic()
